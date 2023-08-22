"""
Script: generateTfvars.py
Author: Monali Tagunde
Date: 15-08-2023
Description: This script accepts input in file input.xlsx having different sheets for every component.
The script generates different tfvars for each component which can be used in terraform to deploy the
infrastructure. Script helps automate the process, faster the infrastructure creation and also reduces
the chances of manual errors.
"""

import csv
import json
import os
import openpyxl
import io


def getCsvdata(sheet):
    # Create a temporary CSV file-like object for the sheet
    csv_data = io.StringIO()
    writer = csv.writer(csv_data)
    # Write each row to the CSV data object
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(row)
    return csv_data


def generateFinalTfvarsFile(intermediate_tfvars_file, final_tfvars_file):
    # Read intermediate tfvars file
    with open(intermediate_tfvars_file, "r") as input_file:
        lines = input_file.readlines()
    # Write final route table tfvars file
    output_lines = []
    for line in lines:
        # Use a loop to find and replace the first two occurrences of double quotes
        if '\\"' in line:
            # for CIDRs
            new_line = line.replace('\\"', '', 2)
        else:
            new_line = line.replace('"', '', 2)
        new_line = new_line.replace(':', ' =', 1)


        # Remove trailing comma. Don't remove comma of multiple items
        if '},' not in new_line:
            new_line = new_line.rstrip(',\n') + '\n'
        output_lines.append(new_line)
    # Write the modified content to the output file
    with open(final_tfvars_file, "w") as output_file:
        output_file.writelines(output_lines)
    print(f'=>>> Completed generating final tfvars file {final_tfvars_file}')


def generateFileNames(sheet_name):
    # Form intermediate and output files
    json_file = os.path.join('intermediate', sheet_name + '.json')
    intermediate_tfvars_file = os.path.join('intermediate', sheet_name + '.tfvars')
    final_tfvars_file = os.path.join('output', 'final-' + sheet_name + '.tfvars')
    return json_file, intermediate_tfvars_file, final_tfvars_file


def getCidrArray(header, row):
    cidr_blocks = []
    cidr_items = row[header].split(',')
    try:
        for cidr_item in cidr_items:
            cidr_blocks.append(f'"{cidr_item}"')
    except ValueError:
        pass
    return cidr_blocks


def getRulesArray(header, row):
    route_rules_array = []
    route_rules_items = row[header].split(',')

    try:
        for route_rules_item in route_rules_items:
            route_rule = {}
            route_rules_key_pairs = route_rules_item.split(';')
            options_array = []

            for pair in route_rules_key_pairs:
                key, value = pair.split("=")
                if key != "options":
                    route_rule[key] = value
                else:
                    options_object = {}
                    option_rule = {}
                    protocol, options_key_pairs = value.split('::')
                    if options_key_pairs != '' and '||' in options_key_pairs:
                        options = options_key_pairs.split("||")
                        for option in options:
                            optionKey, optionVal = option.split("<>")
                            option_rule[optionKey] = optionVal
                        options_array.append(option_rule)
                    options_object[protocol] = options_array
                    route_rule[key] = options_object

            route_rules_array.append(route_rule)

    except ValueError:
        pass
    return route_rules_array


def getTags(header, row):
    tags = {}
    try:
        tags_key_pairs = row[header].split(';')
        for pair in tags_key_pairs:
            key, value = pair.split("=")
            tags[key] = value
    except ValueError:
        pass
    return tags


def generateIntermediateFiles(json_file, intermediate_tfvars_file, sheet_name, resource):
    # Write json file
    with open(json_file, 'w') as jsonfile:
        json.dump(resource, jsonfile, indent=4)

    # Convert route table JSON data to tfvars format
    tfvars_content = sheet_name + " = {\n"
    for key, value in resource.items():
        tfvars_content += f'"{key}": {json.dumps(value, indent=4, separators=(",", ": "))},\n'
    tfvars_content += "}\n"

    # Write intermediate .tfvars file
    with open(intermediate_tfvars_file, "w") as tfvars_file:
        tfvars_file.write(tfvars_content)


def processRouteTable(sheet_name, sheet):
    print(f'Processing sheet {sheet_name}')
    # get intermediate and output file names
    json_file, intermediate_tfvars_file, final_tfvars_file = generateFileNames(sheet_name)
    # Generate csv data from excel sheet
    csv_data = getCsvdata(sheet)
    # Reset the CSV data object's position
    csv_data.seek(0)
    # Use DictReader to read and print CSV data
    reader = csv.DictReader(csv_data)
    route_tables = {}
    for row in reader:
        for header, value in row.items():
            if header == 'route_table_name':
                route_table_name = row['route_table_name']
                # print(f'route_table_name : {route_table_name}')
                route_tables[row['route_table_name']] = {}
            elif header == 'compartment_id':
                route_tables[row['route_table_name']]['compartment_id'] = row['compartment_id']
            elif header == 'vcn_id':
                route_tables[row['route_table_name']]['vcn_id'] = row['vcn_id']
            elif header == 'display_name':
                route_tables[row['route_table_name']]['display_name'] = row['display_name']
            elif header == 'route_rules_drg':
                route_rules_drg_array = getRulesArray(header, row)
            elif header == 'route_rules_igw':
                route_rules_igw_array = getRulesArray(header, row)
            elif header == 'route_rules_sgw':
                route_rules_sgw_array = getRulesArray(header, row)
            elif header == 'route_rules_ngw':
                route_rules_ngw_array = getRulesArray(header, row)
            elif header == 'route_rules_lpg':
                route_rules_lpg_array = getRulesArray(header, row)
            elif header == 'route_rules_ip':
                route_rules_ip_array = getRulesArray(header, row)
            elif header == 'freeform_tags':
                freeform_tags = getTags(header, row)
            elif header == 'defined_tags':
                defined_tags = getTags(header, row)

        # Assign rules array objects
        route_tables[row['route_table_name']]['route_rules_drg'] = route_rules_drg_array
        route_tables[row['route_table_name']]['route_rules_igw'] = route_rules_igw_array
        route_tables[row['route_table_name']]['route_rules_sgw'] = route_rules_sgw_array
        route_tables[row['route_table_name']]['route_rules_ngw'] = route_rules_ngw_array
        route_tables[row['route_table_name']]['route_rules_lpg'] = route_rules_lpg_array
        route_tables[row['route_table_name']]['route_rules_ip'] = route_rules_ip_array
        route_tables[row['route_table_name']]['freeform_tags'] = freeform_tags
        route_tables[row['route_table_name']]['defined_tags'] = defined_tags

    # generate intermediate and final files
    generateIntermediateFiles(json_file, intermediate_tfvars_file, sheet_name, route_tables)
    generateFinalTfvarsFile(intermediate_tfvars_file, final_tfvars_file)


def processVcns(sheet_name, sheet):
    print(f'Processing sheet {sheet_name}')
    # get intermediate and output file names
    json_file, intermediate_tfvars_file, final_tfvars_file = generateFileNames(sheet_name)
    # Generate csv data from excel sheet
    csv_data = getCsvdata(sheet)
    # Reset the CSV data object's position
    csv_data.seek(0)
    # Use DictReader to read and print CSV data
    reader = csv.DictReader(csv_data)
    vcns = {}
    for row in reader:
        for header, value in row.items():
            if header == 'vcn_name':
                vcn_name = row['vcn_name']
                # print(f'vcn_name : {vcn_name}')
                vcns[row['vcn_name']] = {}
            elif header == 'compartment_id':
                vcns[row['vcn_name']]['compartment_id'] = row['compartment_id']
            elif header == 'display_name':
                vcns[row['vcn_name']]['display_name'] = row['display_name']
            elif header == 'dns_label':
                vcns[row['vcn_name']]['dns_label'] = row['dns_label']
            elif header == 'cidr_blocks':
                cidr_blocks = getCidrArray(header, row)

        # # Assign rules array objects
        vcns[row['vcn_name']]['cidr_blocks'] = cidr_blocks

    # generate intermediate and final files
    generateIntermediateFiles(json_file, intermediate_tfvars_file, sheet_name, vcns)
    generateFinalTfvarsFile(intermediate_tfvars_file, final_tfvars_file)


def processDrgAttachments(sheet_name, sheet):
    print(f'Processing sheet {sheet_name}')
    # get intermediate and output file names
    json_file, intermediate_tfvars_file, final_tfvars_file = generateFileNames(sheet_name)
    # Generate csv data from excel sheet
    csv_data = getCsvdata(sheet)
    # Reset the CSV data object's position
    csv_data.seek(0)
    # Use DictReader to read and print CSV data
    reader = csv.DictReader(csv_data)
    drg_attachments = {}
    for row in reader:
        for header, value in row.items():
            if header == 'drg_attachment_name':
                drg_attachment_name = row['drg_attachment_name']
                drg_attachments[row['drg_attachment_name']] = {}
            elif header == 'drg_id':
                drg_attachments[row['drg_attachment_name']]['drg_id'] = row['drg_id']
            elif header == 'display_name':
                drg_attachments[row['drg_attachment_name']]['display_name'] = row['display_name']
            elif header == 'drg_route_table_id':
                drg_attachments[row['drg_attachment_name']]['drg_route_table_id'] = row['drg_route_table_id']
            elif header == 'network_details':
                network_details_array = getRulesArray(header, row)
            elif header == 'vcn_id':
                drg_attachments[row['drg_attachment_name']]['vcn_id'] = row['vcn_id']
            elif header == 'freeform_tags':
                freeform_tags = getTags(header, row)
            elif header == 'defined_tags':
                defined_tags = getTags(header, row)

        # Assign rules array objects
        drg_attachments[row['drg_attachment_name']]['network_details'] = network_details_array
        drg_attachments[row['drg_attachment_name']]['freeform_tags'] = freeform_tags
        drg_attachments[row['drg_attachment_name']]['defined_tags'] = defined_tags

    # generate intermediate and final files
    generateIntermediateFiles(json_file, intermediate_tfvars_file, sheet_name, drg_attachments)
    generateFinalTfvarsFile(intermediate_tfvars_file, final_tfvars_file)


def processSecLists(sheet_name, sheet):
    print(f'Processing sheet {sheet_name}')
    # get intermediate and output file names
    json_file, intermediate_tfvars_file, final_tfvars_file = generateFileNames(sheet_name)
    # Generate csv data from excel sheet
    csv_data = getCsvdata(sheet)
    # Reset the CSV data object's position
    csv_data.seek(0)
    # Use DictReader to read and print CSV data
    reader = csv.DictReader(csv_data)
    seclists = {}
    for row in reader:
        for header, value in row.items():
            if header == 'seclist_name':
                seclist_name = row['seclist_name']
                seclists[row['seclist_name']] = {}
            elif header == 'compartment_id':
                seclists[row['seclist_name']]['compartment_id'] = row['compartment_id']
            elif header == 'vcn_id':
                seclists[row['seclist_name']]['vcn_id'] = row['vcn_id']
            elif header == 'display_name':
                seclists[row['seclist_name']]['display_name'] = row['display_name']
            elif header == 'ingress_sec_rules':
                ingress_sec_rules_array = getRulesArray(header, row)
            elif header == 'egress_sec_rules':
                egress_sec_rules_array = getRulesArray(header, row)
            elif header == 'freeform_tags':
                freeform_tags = getTags(header, row)
            elif header == 'defined_tags':
                defined_tags = getTags(header, row)

        # Assign rules array objects
        seclists[row['seclist_name']]['ingress_sec_rules'] = ingress_sec_rules_array
        seclists[row['seclist_name']]['egress_sec_rules'] = egress_sec_rules_array
        seclists[row['seclist_name']]['freeform_tags'] = freeform_tags
        seclists[row['seclist_name']]['defined_tags'] = defined_tags

    # generate intermediate and final files
    generateIntermediateFiles(json_file, intermediate_tfvars_file, sheet_name, seclists)
    generateFinalTfvarsFile(intermediate_tfvars_file, final_tfvars_file)


#adding function for subnets
def processSubnets(sheet_name, sheet):
    print(f'Processing sheet {sheet_name}')
    # get intermediate and output file names
    json_file, intermediate_tfvars_file, final_tfvars_file = generateFileNames(sheet_name)
    # Generate csv data from excel sheet
    csv_data = getCsvdata(sheet)
    # Reset the CSV data object's position
    csv_data.seek(0)
    # Use DictReader to read and print CSV data
    reader = csv.DictReader(csv_data)
    subnets = {}
    for row in reader:
        for header, value in row.items():
            if header == 'subnet_name':
                subnet_name = row['subnet_name']
                subnets[row['subnet_name']] = {}
            elif header == 'availability_domain':
                subnets[row['subnet_name']]['availability_domain'] = row['availability_domain']
            elif header == 'cidr_block':
                subnets[row['subnet_name']]['cidr_block'] = row['cidr_block']
            elif header == 'compartment_id':
                subnets[row['subnet_name']]['compartment_id'] = row['compartment_id']
            elif header == 'vcn_id':
                subnets[row['subnet_name']]['vcn_id'] = row['vcn_id']
            elif header == 'display_name':
                subnets[row['subnet_name']]['display_name'] = row['display_name']
            elif header == 'prohibit_public_ip_on_vnic':
                subnets[row['subnet_name']]['prohibit_public_ip_on_vnic'] = row['prohibit_public_ip_on_vnic'].lower()
            elif header == 'route_table_id':
                subnets[row['subnet_name']]['route_table_id'] = row['route_table_id']
            elif header == 'dns_label':
                subnets[row['subnet_name']]['dns_label'] = row['dns_label']
            elif header == 'dhcp_options_id':
                subnets[row['subnet_name']]['dhcp_options_id'] = row['dhcp_options_id']
            elif header == 'security_list_ids':
                security_list_ids = row['security_list_ids'].split(',')
                subnets[row['subnet_name']]['security_list_ids'] = security_list_ids
            elif header == 'freeform_tags':
                freeform_tags = getTags(header, row)
            elif header == 'defined_tags':
                defined_tags = getTags(header, row)

        # Assign rules array objects
        subnets[row['subnet_name']]['freeform_tags'] = freeform_tags
        subnets[row['subnet_name']]['defined_tags'] = defined_tags

    # generate intermediate and final files
    generateIntermediateFiles(json_file, intermediate_tfvars_file, sheet_name, subnets)
    generateFinalTfvarsFile(intermediate_tfvars_file, final_tfvars_file)

#Function for Virtual Machine tfvars file
def processInstances(sheet_name, sheet):
    print(f'Processing sheet {sheet_name}')
    # get intermediate and output file names
    json_file, intermediate_tfvars_file, final_tfvars_file = generateFileNames(sheet_name)
    # Generate csv data from excel sheet
    csv_data = getCsvdata(sheet)
    # Reset the CSV data object's position
    csv_data.seek(0)
    # Use DictReader to read and print CSV data
    reader = csv.DictReader(csv_data)
    instances = {}

    for row in reader:
        for header, value in row.items():
            if header == 'instance_name':
                instance_name = row['instance_name']
                instances[row['instance_name']] = {}
            elif header == 'availability_domain':
                instances[row['instance_name']]['availability_domain'] = int(row['availability_domain'])
            elif header == 'compartment_id':
                instances[row['instance_name']]['compartment_id'] = row['compartment_id']
            elif header == 'shape':
                instances[row['instance_name']]['shape'] = row['shape']
            elif header == 'display_name':
                instances[row['instance_name']]['display_name'] = row['display_name']
            elif header == 'boot_volume_size_in_gbs':
                instances[row['instance_name']]['boot_volume_size_in_gbs'] = int(row['boot_volume_size_in_gbs'])
            elif header == 'fault_domain':
                instances[row['instance_name']]['fault_domain'] = row['fault_domain']
            elif header == 'source_id':
                instances[row['instance_name']]['source_id'] = row['source_id']
            elif header == 'source_type':
                instances[row['instance_name']]['source_type'] = row['source_type']
            elif header == 'network_compartment_id':
                instances[row['instance_name']]['network_compartment_id'] = row['network_compartment_id']
            elif header == 'vcn_compartment_id':
                instances[row['instance_name']]['vcn_compartment_id'] = row['vcn_compartment_id']
            elif header == 'vcn_name':
                instances[row['instance_name']]['vcn_name'] = row['vcn_name']
            elif header == 'subnet_id':
                instances[row['instance_name']]['subnet_id'] = row['subnet_id']
            elif header == 'assign_public_ip':
                if row['assign_public_ip'].lower() == "true":
                    boolean_value = True
                elif row['assign_public_ip'].lower().lower() == "false":
                    boolean_value = False
                instances[row['instance_name']]['assign_public_ip'] = boolean_value
            elif header == 'private_ip':
                instances[row['instance_name']]['private_ip'] = row['private_ip']
            elif header == 'ocpus':
                instances[row['instance_name']]['ocpus'] = row['ocpus']
            elif header == 'memory_in_gbs':
                instances[row['instance_name']]['memory_in_gbs'] = int(row['memory_in_gbs'])
            elif header == 'update_is_pv_encryption_in_transit_enabled':
                if row['update_is_pv_encryption_in_transit_enabled'].lower() == "true":
                    boolean_value = True
                elif row['update_is_pv_encryption_in_transit_enabled'].lower().lower() == "false":
                    boolean_value = False
                instances[row['instance_name']]['update_is_pv_encryption_in_transit_enabled'] = boolean_value
            elif header == 'freeform_tags':
                freeform_tags = getTags(header, row)
            elif header == 'defined_tags':
                defined_tags = getTags(header, row)

        # Assign rules array objects
        instances[row['instance_name']]['freeform_tags'] = freeform_tags
        instances[row['instance_name']]['defined_tags'] = defined_tags

        # generate intermediate and final files
        generateIntermediateFiles(json_file, intermediate_tfvars_file, sheet_name, instances)
        generateFinalTfvarsFile(intermediate_tfvars_file, final_tfvars_file)




###################################################################################################
#####                                generarteTfvars,py script                                #####
###################################################################################################

myTfAutomationDir = "C:/workdir/PycharmProjects/Tutorials/tf-automation"
os.chdir(myTfAutomationDir)
current_directory = os.getcwd()
print(f'current_directory : {current_directory}')
intermediate_dir = os.path.join(current_directory, 'intermediate')
output_dir = os.path.join(current_directory, 'output')
if not os.path.exists(intermediate_dir):
    os.makedirs(intermediate_dir)
if not os.path.exists(output_dir):
    os.makedirs(output_dir)


# Create a dictionary mapping sheet names to functions
sheet_function_mapping = {
    'route_tables': processRouteTable,
    'vcns': processVcns,
    'drg_attachments': processDrgAttachments,
    'seclists': processSecLists,
    'subnets': processSubnets,
    'instances': processInstances,
}

# Open the Excel file
file_path = 'input.xlsx'
workbook = openpyxl.load_workbook(file_path, data_only=True)

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    if sheet_name in sheet_function_mapping:
        sheet_function = sheet_function_mapping[sheet_name]
        sheet = workbook[sheet_name]
        sheet_function(sheet_name, sheet)
    else:
        print(f"No function found for sheet: {sheet_name}")


# Close the Excel file
workbook.close()